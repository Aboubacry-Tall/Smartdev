from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import csv
import io
import logging
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class ImportTransferLinesWizard(models.TransientModel):
    _name = 'import.transfer.lines.wizard'
    _description = 'Assistant d\'import de lignes de virement'

    transfer_id = fields.Many2one(
        'sogem.transfer', string='Virement', required=True,
        default=lambda self: self.env.context.get('active_id'),readonly=True)
    file = fields.Binary(string='Fichier à importer', required=True)
    filename = fields.Char(string='Nom du fichier')

    def action_import(self):
        if not self.file or not self.filename:
            raise UserError(_('Veuillez sélectionner un fichier à importer.'))
        if self.filename.lower().endswith('.csv'):
            self._import_csv()
        elif self.filename.lower().endswith('.xlsx'):
            self._import_xlsx()
        else:
            raise UserError(
                _('Format de fichier non supporté. Veuillez utiliser un fichier .csv ou .xlsx.'))

    def _import_csv(self):
        data = base64.b64decode(self.file)
        file_stream = io.StringIO(data.decode('utf-8'))
        reader = csv.DictReader(file_stream)
        self._process_rows(reader)

    def _import_xlsx(self):
        if not load_workbook:
            raise UserError(
                _('Le module openpyxl est requis pour lire les fichiers .xlsx.'))
        data = base64.b64decode(self.file)
        file_stream = io.BytesIO(data)
        workbook = load_workbook(file_stream, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in next(sheet.rows)]
        rows = []
        for row in sheet.iter_rows(min_row=2):
            row_dict = {}
            for header, cell in zip(headers, row):
                if header:  # Only process cells that have a header
                    value = cell.value
                    if value is not None:  # Convert cell value to string if it's not None
                        value = str(value).strip()
                    row_dict[header] = value
            if any(row_dict.values()):  # Only append non-empty rows
                rows.append(row_dict)
        self._process_rows(rows)

    def _process_rows(self, rows):
        TransferLine = self.env['sogem.transfer.line']
        Beneficiary = self.env['sogem.beneficiary']
        created = 0
        errors = []
        
        _logger = logging.getLogger(__name__)
        _logger.info(f"Starting import with {len(rows)} rows")
        
        for row_index, row in enumerate(rows, start=1):
            _logger.info(f"Processing row {row_index}: {row}")
            try:
                # Get raw matricule and clean it
                matricule_raw = str(row.get('Matricule', '')).strip()
                _logger.info(f"Raw matricule: {matricule_raw}")
                if not matricule_raw:
                    errors.append(f"Ligne {row_index}: Matricule manquant")
                    continue
                
                # Remove any decimal part for matricule
                if '.' in matricule_raw or ',' in matricule_raw:
                    matricule_raw = matricule_raw.split('.')[0].split(',')[0]
                    
                # Convert to integer
                try:
                    matricule = int(matricule_raw)
                    _logger.info(f"Converted matricule: {matricule}")
                except (ValueError, TypeError) as e:
                    _logger.error(f"Error converting matricule: {str(e)}")
                    errors.append(f"Ligne {row_index}: Le matricule '{matricule_raw}' n'est pas un nombre valide")
                    continue
                
                montant = row.get('Montant')
                _logger.info(f"Raw montant: {montant}")
                if not montant:
                    errors.append(f"Ligne {row_index}: Montant manquant")
                    continue
                    
                try:
                    montant = float(str(montant).replace(',', '.'))
                    _logger.info(f"Converted montant: {montant}")
                except (ValueError, TypeError) as e:
                    _logger.error(f"Error converting montant: {str(e)}")
                    errors.append(f"Ligne {row_index}: Le montant '{montant}' n'est pas un nombre valide")
                    continue
                
                beneficiary = Beneficiary.search([('matricule', '=', matricule)], limit=1)
                _logger.info(f"Found beneficiary: {beneficiary}")
                if not beneficiary:
                    errors.append(f"Ligne {row_index}: Bénéficiaire non trouvé pour le matricule {matricule}")
                    continue

                # Check if beneficiary has a bank account
                if not beneficiary.bank_account_ids:
                    errors.append(f"Ligne {row_index}: Le bénéficiaire (matricule {matricule}) n'a pas de compte bancaire")
                    continue

                account_id = beneficiary.bank_account_ids[:1].id
                _logger.info(f"Found account_id: {account_id}")
                if not account_id:
                    errors.append(f"Ligne {row_index}: Impossible d'accéder au compte bancaire du bénéficiaire (matricule {matricule})")
                    continue
                    
                transfer_line = TransferLine.create({
                    'transfer_id': self.transfer_id.id,
                    'beneficiary_id': beneficiary.id,
                    'amount': montant,
                    'account_id': account_id,
                })
                _logger.info(f"Created transfer line: {transfer_line}")
                created += 1
                
            except Exception as e:
                _logger.error(f"Error processing row {row_index}: {str(e)}")
                errors.append(f"Ligne {row_index}: Erreur lors de la création de la ligne: {str(e)}")
                self.env.cr.rollback()
                self.env.cr.begin()
        
        _logger.info(f"Import finished. Created: {created}, Errors: {len(errors)}")
        
        # If no lines were created at all, raise an error
        if not created:
            error_message = "Aucune ligne de virement n'a été créée.\n\nDétails des erreurs:\n" + "\n".join(errors)
            raise UserError(_(error_message))
            
        # If we have some successes and some errors, show the warning but keep the successful imports
        if errors:
            message = _('Import terminé.\nLignes créées: %d\nLignes en erreur: %d\n\nDétails des erreurs:\n%s') % (
                created, len(errors), '\n'.join(errors))
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Résultat de l\'import'),
                    'message': message,
                    'type': 'warning',
                    'sticky': True,
                }
            }
        
        # All successful
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Succès'),
                'message': _('Import terminé avec succès. %d lignes créées.') % created,
                'type': 'success',
            }
        }
